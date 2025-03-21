from openpyxl import Workbook, load_workbook
import os

class DBMS:
    def __init__(self, file_name="database.xlsx", sheet_name="Records"):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self._initialize_database()

    def _initialize_database(self):
        """Creates an Excel file with a default sheet if it does not exist."""
        if not os.path.exists(self.file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            wb.save(self.file_name)

    def _get_workbook(self):
        """Loads the Excel workbook."""
        return load_workbook(self.file_name)

    def set_headers(self, headers):
        """Sets column headers if the sheet is empty."""
        wb = self._get_workbook()
        ws = wb[self.sheet_name]
        if ws.max_row == 1:  # Only set headers if the sheet is empty
            ws.append(headers)
            wb.save(self.file_name)

    def add_record(self, record):
        """Adds a new record to the sheet."""
        wb = self._get_workbook()
        ws = wb[self.sheet_name]
        ws.append(record)
        wb.save(self.file_name)

    def fetch_records(self):
        """Returns all records as a list of tuples."""
        wb = self._get_workbook()
        ws = wb[self.sheet_name]
        return [tuple(cell.value for cell in row) for row in ws.iter_rows(values_only=True)]

    def update_record(self, search_column, search_value, update_column, new_value):
        """Updates a record in the sheet."""
        wb = self._get_workbook()
        ws = wb[self.sheet_name]
        headers = [cell.value for cell in ws[1]]
        if search_column not in headers or update_column not in headers:
            return

        search_idx = headers.index(search_column)
        update_idx = headers.index(update_column)

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[search_idx].value == search_value:
                row[update_idx].value = new_value

        wb.save(self.file_name)

    def sort_records(self, sort_column, ascending=True):
        """Sorts the sheet based on a specified column."""
        wb = self._get_workbook()
        ws = wb[self.sheet_name]
        headers = [cell.value for cell in ws[1]]

        if sort_column not in headers:
            return

        sort_idx = headers.index(sort_column)

        # Extract data excluding headers
        data = [row for row in ws.iter_rows(values_only=True)]
        headers, records = data[0], data[1:]  # Separate headers from records

        # Sort records
        records.sort(key=lambda x: x[sort_idx], reverse=not ascending)

        # Clear existing records and re-add sorted data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        for row in records:
            ws.append(row)

        wb.save(self.file_name)
